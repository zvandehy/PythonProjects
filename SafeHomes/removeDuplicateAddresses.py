import openpyxl as xl

wb = xl.load_workbook("CRM_Records_COPY.xlsx")
combined = wb.create_sheet("Combined")
sheets = wb.sheetnames
print(sheets)
for i,sheet in enumerate(sheets):
    print(i, sheet)
    ws = wb[sheet]
    new_sheet = wb.create_sheet(sheet + " - Reduced")
    
    addresses = set()
    names = set()

    for row in ws.iter_rows(max_row=5500, values_only=True):
        
        address = row[4]
        if row[1] is None or row[1] == "":
            name = row[3]
        else:
            name = row[1].strip() + row[2].strip()

        if address in addresses or name in names:
            continue
        else: 
            new_sheet.append(row)
            if i == 0 or i == 2 or i==4:
                combined.append(row)
            addresses.add(row[4])
            names.add(name)
    
    wb.save("CRM_Records_COPY_REDUCED.xlsx")